from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'


# CARREGANDO O ARQUIVO
workbook: Workbook = load_workbook(WORKBOOK_PATH)



# worksheet: Worksheet = workbook.active # comentar


# # Nome para a planilha
sheet_name = 'Sheet'

# # Criamos a planilha
# workbook.create_sheet(sheet_name, 0)

# # Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# # Remover uma planilha
# workbook.remove(workbook['Sheet'])


# LER 
for row in worksheet.iter_rows():
    for collun in row:
        print(collun.value, end='\t')
    print()

# ALTERAR
worksheet['B3'].value = 14





workbook.save(WORKBOOK_PATH)